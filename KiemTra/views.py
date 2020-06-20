from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from .models import Lop, Mon, BaiKiemTra, ThanhTich
import openpyxl, os
from random import shuffle, sample
from Chinh.models import ThanhVien

# Create your views here.
def layDSMon(request, idlop):
    lp = get_object_or_404(Lop, pk=idlop)
    return render(request, 'KiemTra/ChaoMungLop.html', {'lop': lp})

def layDSBaiKT(request, idlop, idmon):
    mhs = get_object_or_404(Mon, pk=idmon)
    dsbai = mhs.baikiemtra_set.all()
    return render(request, 'KiemTra/DSBaiKT.html', {'mh': dsbai})

def docFileKiemTra(idb):
    bai = get_object_or_404(BaiKiemTra, pk=idb)

    pathfile=bai.NoiDung.file

    wb_obj = openpyxl.load_workbook(pathfile)
    sheet_obj = wb_obj.active
    max_cot = sheet_obj.max_column
    max_dong = sheet_obj.max_row
    chs = []
    for d in range(2, max_dong+1):
        chsdetail = {}
        chtraloi = {}
        for c in range(1, max_cot+1):
            chsdetail['stt']=d-1
            cell_obj = sheet_obj.cell(row=d,column=c)
            if c == 1:
                chsdetail['cauhoi'] = cell_obj.value
            elif c == 2:
                chtraloi['a'] = cell_obj.value
            elif c == 3:
                chtraloi['b'] = cell_obj.value
            elif c == 4:
                chtraloi['c'] = cell_obj.value
            elif c == 5:
                chtraloi['d'] = cell_obj.value
            elif c == 6:
                chsdetail['dapan'] = cell_obj.value
            else:
                chsdetail['hinhbt'] = cell_obj.value
                chsdetail['hinh'] = cell_obj.value
                if chsdetail['hinh'] is not None:
                    for hinh in bai.hinhlienquanbaikt_set.all():
                        if chsdetail['hinh'] in str(hinh.HinhLienQuan):
                            chsdetail['hinh'] = hinh.HinhLienQuan
                            break
        chsdetail['traloi'] = chtraloi
        chs.append(chsdetail)
    return chs
def xaoTronCauHoiVaTraLoi(idbai):
    chs = docFileKiemTra(idbai)
    #shuffle(chs)
    t = len(chs)
    chs_new = sample(chs, t)
    for ch in chs_new:
        d = ch['traloi']
        keys = list(d.keys())
        #shuffle(keys)
        keys_new=sample(keys,len(keys))
        shu_d = dict()
        for k in keys_new:
            #shu_d.update({k:d[k]})
            shu_d[k]=d[k]
        ch['traloi'] = {}
        ch['traloi'] = shu_d
    return chs_new

def xuLyFormCauHoi(request, idlop, idmon, idbai):
    chs_sauxao = xaoTronCauHoiVaTraLoi(idbai)
    tongch = len(chs_sauxao)
    socaudung = 0
    chs_kq = chs_sauxao

    bai = get_object_or_404(BaiKiemTra, pk=idbai)
    #pathbai =str(bai.NoiDung)
    #tenbai = pathbai.split("/")[-1]
    #pathfilekq = "BaiKT/"+pathbai.split(tenbai)[0]

    pathfilekq=os.path.dirname(str(bai.NoiDung.url))
    filekq = pathfilekq + pathfilekq[0] + request.session['tentk'] + "_BaiLam.xlsx"
    pathfilekq=os.path.join(os.getcwd(),filekq[1:])
    request.session['pathfilekq'] = pathfilekq #duong dan tuyet doi cua file kq
    request.session['filekq'] = filekq #url cua file kq
    request.session['tongch'] = tongch
    
    book = openpyxl.Workbook()
    sheet = book.active
    sheet['A1'] = 'STT'
    sheet['B1'] = 'Cau hoi'
    sheet['C1'] = 'A'
    sheet['D1'] = 'B'
    sheet['E1'] = 'C'
    sheet['F1'] = 'D'
    sheet['G1'] = 'Tra loi'
    sheet['H1'] = 'Dap an'
    sheet['I1'] = 'Hinh'
    sheet['J1'] = 'STT ban dau'
    
    i=2
    for ch in chs_sauxao:
        sheet['A'+str(i)] = i-1
        sheet['B'+str(i)] = ch['cauhoi']
        count = 0
        for key, value in ch['traloi'].items():
            count +=1
            if count == 1:
                sheet['C'+str(i)] = value
            elif count==2:
                sheet['D'+str(i)] = value
            elif count==3:
                sheet['E'+str(i)] = value
            else:
                sheet['F'+str(i)] = value
            if key == "a":
                dapan = value
        if dapan == sheet.cell(row=i,column=3).value:
            sheet['H'+str(i)] ="A"
        elif dapan == sheet.cell(row=i,column=4).value:
            sheet['H'+str(i)] ="B"
        elif dapan == sheet.cell(row=i,column=5).value:
            sheet['H'+str(i)] ="C"
        elif dapan == sheet.cell(row=i,column=6).value:
            sheet['H'+str(i)] ="D"
        else:
            pass
        sheet['I'+str(i)] = ch['hinhbt']
        sheet['J'+str(i)] = ch['stt']

        i = i+1

    book.save(pathfilekq)

    return render(request, 'KiemTra/BaiKT.html',{'chs':chs_sauxao,'tongch':tongch,'idlop':idlop,'idmon':idmon,'idbai':idbai})

def luuThanhTich(idtv,idbai,diem):
    try:
        tt = ThanhTich.objects.get(thanhvien=ThanhVien.objects.get(pk=idtv),baikiemtra=BaiKiemTra.objects.get(id=idbai))
    except ThanhTich.DoesNotExist:
        thanhtich = ThanhTich()
        thanhtich.thanhvien=ThanhVien.objects.get(id=idtv)
        thanhtich.baikiemtra=BaiKiemTra.objects.get(id=idbai)
        thanhtich.Diem=diem
        thanhtich.save()
    else:
        tt.Diem=diem
        tt.save()

def hienThiKQ(request, idlop, idmon, idbai):
    diem = 0
    socaudung=0
    tongch = request.session['tongch']

    if request.method == "POST":
        
        keys = list(request.POST.keys())
        if len(keys) > 2: #neu user khong tra loi cau nao thi chi co request voi key la next tuong ung voi input hidden
            pathfilekq=os.path.join(os.getcwd(),request.session['pathfilekq'])
            
            wb_obj = openpyxl.load_workbook(pathfilekq)
            sheet_obj = wb_obj.active   
            
            for i in range(1,tongch+1):
                k = request.POST[str(i)]
                if k.split("_")[0] == "a":
                    socaudung += 1
                
                for dong in sheet_obj.iter_cols(10):
                    for cell in dong:
                        if cell.value == i:
                            if k.split("_")[1] == "1":
                                sheet_obj.cell(row=cell.row,column=7).value="A"
                            elif k.split("_")[1] == "2":
                                sheet_obj.cell(row=cell.row,column=7).value="B"
                            elif k.split("_")[1] == "3":
                                sheet_obj.cell(row=cell.row,column=7).value="C"
                            else:
                                sheet_obj.cell(row=cell.row,column=7).value="D" 
            wb_obj.save(pathfilekq)    
        diem = socaudung * 10/tongch
        idtv=request.session['idtv']
        luuThanhTich(idtv,idbai,diem)
    return render(request, 'KiemTra/ThongBaoKq.html',{'diem': diem})
